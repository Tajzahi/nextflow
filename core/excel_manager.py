import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from core.config import ExcelColor


class ExcelManager:
    # C-014: Validasi Struktur File
    KOLOM_WAJIB = ["NIK", "Nama", "Tanggal_Lahir"]

    def __init__(self, file_path):
        self.file_path = file_path
        self.df = None
        self.workbook = None # C-012: Buffer saving

    def buka_koneksi(self):
        """Memuat workbook ke memori untuk manipulasi batch (C-012)"""
        if self.workbook is None:
            self.workbook = load_workbook(self.file_path)
        return self.workbook

    def tutup_koneksi(self):
        """Menghapus workbook dari memori"""
        if self.workbook:
            self.workbook.close()
            self.workbook = None

    def simpan_perubahan(self):
        """Melakukan sinkronisasi memori ke disk (Flush)"""
        if self.workbook:
            self.workbook.save(self.file_path)
            print(">>> Perubahan Excel berhasil disimpan ke disk.")

    def baca_data(self):
        self.df = pd.read_excel(self.file_path, engine="openpyxl")
        self.df = self.df.fillna("")  # Mencegah NaN terketik di web
        return self.df

    def validasi_format(self):
        """Mengecek apakah kolom wajib ada dalam file Excel (C-014)"""
        if self.df is None:
            self.baca_data()
        
        missing = [col for col in self.KOLOM_WAJIB if col not in self.df.columns]
        if missing:
            return False, f"Kolom wajib tidak ditemukan: {', '.join(missing)}"
        return True, "Format Valid"

    def cek_apakah_dilewati(self, index_row, workbook_obj=None):
        """
        C-011: Perketat deteksi baris selesai.
        Hanya anggap dilewati jika warnanya presisi sesuai standar.
        """
        try:
            if workbook_obj:
                wb = workbook_obj
            elif self.workbook:
                wb = self.workbook
            else:
                wb = load_workbook(self.file_path, read_only=True)
                
            ws = wb.active
            baris_excel = index_row + 2
            cell = ws.cell(row=baris_excel, column=1)
            
            # LOGIKA TOLERAN: Jika ada 'fill' (latar belakang) dan tipenya bukan None
            # maka kita anggap baris ini sudah diproses.
            if cell.fill and cell.fill.fill_type is not None:
                # 'none' adalah default untuk sel tanpa warna di openpyxl
                if cell.fill.fill_type != 'none':
                    return True
                
            return False
        except Exception:
            return False

    def _warnai_baris(self, index_row, hex_color, log_pesan):
        """C-012: Modifikasi di memori tanpa save langsung (Fast I/O)"""
        try:
            # Gunakan koneksi yang ada atau buka otomatis
            wb = self.buka_koneksi()
            ws = wb.active
            fill = PatternFill(
                start_color=hex_color, end_color=hex_color, fill_type="solid"
            )
            baris_excel = index_row + 2
            for cell in ws[baris_excel]:
                cell.fill = fill
            
            # JANGAN panggil wb.save() disini. Biarkan Controller yang memanggil simpan_perubahan()
            print(log_pesan)
        except Exception as e:
            print(f"Gagal mewarnai Excel: {e}")

    def tandai_sukses(self, index_row):
        self._warnai_baris(
            index_row,
            ExcelColor.HIJAU,
            f">>> Baris {index_row + 2} diwarnai HIJAU di memori.",
        )

    def tandai_error(self, index_row):
        self._warnai_baris(
            index_row,
            ExcelColor.MERAH,
            f"!!! Baris {index_row + 2} ditandai MERAH di memori.",
        )

    def tandai_kuning(self, index_row):
        self._warnai_baris(
            index_row,
            ExcelColor.KUNING,
            f">>> Baris {index_row + 2} ditandai KUNING di memori.",
        )

    def tandai_sistem_error(self, index_row):
        self._warnai_baris(
            index_row,
            ExcelColor.BIRU,
            f"!!! Baris {index_row + 2} ditandai BIRU di memori.",
        )
