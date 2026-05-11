import sqlite3
import os
import sys
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import base64
import logging
import json
from datetime import datetime, timedelta
from typing import Optional, TYPE_CHECKING
from core.constants import APP_NAME, DB_FILE_NAME

if TYPE_CHECKING:
    from core.database_manager import DatabaseManager

# Security Imports
try:
    import keyring
    KEYRING_AVAILABLE = True
except ImportError:
    KEYRING_AVAILABLE = False

try:
    from cryptography.fernet import Fernet
    FERNET_AVAILABLE = True
except ImportError:
    FERNET_AVAILABLE = False

logger = logging.getLogger(__name__)

# Constants untuk Security
SERVICE_NAME = "NextFlowPro"
TOKEN_KEY = "license_token"
DEVICE_ID_KEY = "device_id"

def get_app_data_dir() -> Path:
    """
    Mendapatkan direktori AppData dengan fallback aman.
    Returns Path object.
    """
    # Prioritas 1: APPDATA (normal Windows)
    app_data = os.getenv('APPDATA')
    if app_data and os.path.exists(app_data):
        return Path(app_data) / APP_NAME

    # Prioritas 2: LOCALAPPDATA
    local_app_data = os.getenv('LOCALAPPDATA')
    if local_app_data and os.path.exists(local_app_data):
        return Path(local_app_data) / APP_NAME

    # Prioritas 3: Home directory
    home = os.path.expanduser("~")
    if os.path.exists(home):
        return Path(home) / f".{APP_NAME.lower()}"

    # Prioritas 4 (last resort): folder di samping .exe
    base_path = os.path.dirname(os.path.abspath(sys.executable
            if getattr(sys, 'frozen', False) else __file__))
    return Path(base_path) / "data"

class DatabaseManager:
    def __init__(self):
        # 1. Menentukan Lokasi Folder AppData dengan robust logic
        self.app_dir = get_app_data_dir()
        self.app_dir.mkdir(parents=True, exist_ok=True)
            
        self.db_path = self.app_dir / DB_FILE_NAME
        self.key_file = self.app_dir / ".fernet_key"
        
        self.buat_tabel()

    def _get_connection(self):
        return sqlite3.connect(str(self.db_path), timeout=20)

    def buat_tabel(self):
        """Membuat struktur database jika belum pernah dibuat"""
        conn = self._get_connection()
        cursor = conn.cursor()
        
        # Tabel Log Eksekusi
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS log_laporan (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                tanggal TEXT NOT NULL,
                baris INTEGER NOT NULL,
                status TEXT NOT NULL,
                keterangan TEXT NOT NULL,
                is_synced INTEGER DEFAULT 1,
                waktu_eksekusi TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(tanggal, baris)
            )
        ''')

        # Tabel Pengaturan (Sekarang mendukung metadata enkripsi & cache device_id)
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS pengaturan (
                kunci TEXT PRIMARY KEY,
                nilai TEXT NOT NULL,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        conn.commit()

        # --- MIGRATION: Tambah kolom is_synced jika belum ada (TASK Deferred Sync) ---
        try:
            cursor.execute("ALTER TABLE log_laporan ADD COLUMN is_synced INTEGER DEFAULT 1")
            conn.commit()
        except Exception as e:
            logger.debug(f"Kolom is_synced mungkin sudah ada: {e}")
            pass

        conn.close()

    # =========================================================================
    # TOKEN STORAGE - ENCRYPTED (TASK W1)
    # =========================================================================

    def simpan_token(self, token: str) -> bool:
        """
        Simpan token dengan security berlapis:
        1. Primary: Windows Credential Manager (keyring)
        2. Fallback: Fernet symmetric encryption
        3. Legacy: Base64 encoding
        """
        if not token or not token.strip():
            return False
        
        token = token.strip()
        
        # METHOD 1: KEYRING (Windows Credential Manager)
        if KEYRING_AVAILABLE:
            try:
                keyring.set_password(SERVICE_NAME, TOKEN_KEY, token)
                logger.info("[OK] Token tersimpan di Windows Credential Manager")
                # Hapus token dari DB agar tidak double storage (lebih aman)
                self._hapus_db_only_token()
                return True
            except Exception as e:
                logger.warning(f"[WARN] Keyring gagal: {e}. Fallback ke Fernet.")

        # METHOD 2: FERNET ENCRYPTION (Fallback)
        if FERNET_AVAILABLE:
            try:
                key = self._get_or_create_fernet_key()
                cipher = Fernet(key)
                encrypted_token = cipher.encrypt(token.encode()).decode('utf-8')
                
                conn = self._get_connection()
                cursor = conn.cursor()
                cursor.execute("INSERT OR REPLACE INTO pengaturan (kunci, nilai) VALUES (?, ?)", 
                             ('token_encrypted', encrypted_token))
                conn.commit()
                conn.close()
                logger.info("[OK] Token tersimpan dengan Fernet encryption")
                return True
            except Exception as e:
                logger.error(f"[ERROR] Fernet encryption gagal: {e}")

        # METHOD 3: BASE64 (Legacy/Last Resort)
        try:
            encoded_token = base64.b64encode(token.encode()).decode('utf-8')
            conn = self._get_connection()
            cursor = conn.cursor()
            cursor.execute("INSERT OR REPLACE INTO pengaturan (kunci, nilai) VALUES (?, ?)", 
                         ("AUTH_TOKEN", encoded_token))
            conn.commit()
            conn.close()
            logger.warning("⚠️ Token tersimpan dengan Base64 (TIDAK AMAN)")
            return True
        except Exception as e:
            logger.error(f"❌ Gagal simpan token: {e}")
            return False

    def ambil_token(self) -> Optional[str]:
        """Ambil token dari storage (Priority: keyring > fernet > base64)"""
        # METHOD 1: KEYRING
        if KEYRING_AVAILABLE:
            try:
                token = keyring.get_password(SERVICE_NAME, TOKEN_KEY)
                if token:
                    return token
            except Exception as e:
                logger.debug(f"Error ambil keyring: {e}")

        # METHOD 2: FERNET
        if FERNET_AVAILABLE:
            try:
                conn = self._get_connection()
                cursor = conn.cursor()
                cursor.execute("SELECT nilai FROM pengaturan WHERE kunci = ?", ('token_encrypted',))
                row = cursor.fetchone()
                conn.close()
                
                if row:
                    key = self._get_or_create_fernet_key()
                    cipher = Fernet(key)
                    return cipher.decrypt(row[0].encode()).decode('utf-8')
            except Exception as e:
                logger.debug(f"Error ambil fernet: {e}")

        # METHOD 3: BASE64 (Legacy)
        try:
            conn = self._get_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT nilai FROM pengaturan WHERE kunci = ?", ("AUTH_TOKEN",))
            row = cursor.fetchone()
            conn.close()
            
            if row and row[0]:
                try:
                    return base64.b64decode(row[0].encode()).decode()
                except Exception:
                    return row[0] # Plain text fallback
        except Exception:
            pass
            
        return None

    def hapus_token(self) -> bool:
        """Hapus token dari semua storage"""
        success = True
        if KEYRING_AVAILABLE:
            try:
                keyring.delete_password(SERVICE_NAME, TOKEN_KEY)
            except Exception:
                pass
        
        try:
            conn = self._get_connection()
            cursor = conn.cursor()
            cursor.execute("DELETE FROM pengaturan WHERE kunci IN ('AUTH_TOKEN', 'token_encrypted')")
            conn.commit()
            conn.close()
        except Exception:
            success = False
        return success

    def _hapus_db_only_token(self):
        try:
            conn = self._get_connection()
            cursor = conn.cursor()
            cursor.execute("DELETE FROM pengaturan WHERE kunci IN ('AUTH_TOKEN', 'token_encrypted')")
            conn.commit()
            conn.close()
        except Exception as e:
            logger.error(f"Gagal menghapus token lama di DB: {e}")
            pass

    def _get_or_create_fernet_key(self) -> bytes:
        if self.key_file.exists():
            return self.key_file.read_bytes()
        
        key = Fernet.generate_key()
        self.key_file.write_bytes(key)
        # Windows Hidden Attribute
        try:
            import ctypes
            ctypes.windll.kernel32.SetFileAttributesW(str(self.key_file), 0x02)
        except Exception as e:
            logger.debug(f"Gagal set atribut file tersembunyi: {e}")
            pass
        return key

    # =========================================================================
    # DEVICE ID CACHE (TASK W2)
    # =========================================================================

    def simpan_device_id(self, device_id: str):
        try:
            conn = self._get_connection()
            cursor = conn.cursor()
            cursor.execute("INSERT OR REPLACE INTO pengaturan (kunci, nilai) VALUES (?, ?)", 
                         (DEVICE_ID_KEY, device_id))
            conn.commit()
            conn.close()
        except Exception as e:
            logger.error(f"Gagal cache device_id: {e}")

    def ambil_device_id(self) -> Optional[str]:
        try:
            conn = self._get_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT nilai FROM pengaturan WHERE kunci = ?", (DEVICE_ID_KEY,))
            row = cursor.fetchone()
            conn.close()
            return row[0] if row else None
        except Exception as e:
            logger.debug(f"Device ID belum tercache: {e}")
            return None

    # =========================================================================
    # STATUS CACHE (TASK W5 - OFFLINE GRACE PERIOD)
    # =========================================================================

    def simpan_last_status(self, status_data: dict, ttl_minutes: int = 60):
        """Simpan status user terakhir dengan masa berlaku (TTL)."""
        try:
            cache_data = {
                'status': status_data,
                'cached_at': datetime.now().isoformat(),
                'expires_at': (datetime.now() + timedelta(minutes=ttl_minutes)).isoformat()
            }
            json_str = json.dumps(cache_data)
            conn = self._get_connection()
            cursor = conn.cursor()
            cursor.execute("INSERT OR REPLACE INTO pengaturan (kunci, nilai) VALUES (?, ?)", 
                         ('last_status_cache', json_str))
            conn.commit()
            conn.close()
            logger.debug(f"Status cached successfully (TTL: {ttl_minutes}m)")
        except Exception as e:
            logger.error(f"Gagal simpan status cache: {e}")

    def ambil_last_status(self) -> Optional[dict]:
        """Ambil status cache jika belum kedaluwarsa."""
        try:
            conn = self._get_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT nilai FROM pengaturan WHERE kunci = ?", ('last_status_cache',))
            row = cursor.fetchone()
            conn.close()
            
            if not row:
                return None
            
            data = json.loads(row[0])
            expires_at = datetime.fromisoformat(data['expires_at'])
            
            if datetime.now() > expires_at:
                logger.debug("Status cache expired")
                return None
                
            return data['status']
        except Exception as e:
            logger.error(f"Gagal ambil status cache: {e}")
            return None

    # =========================================================================
    # LOGGING & EXPORT METHODS
    # =========================================================================

    def tambah_log(self, tanggal, baris, status, keterangan, is_synced=1):
        """Tambah log eksekusi dengan status sinkronisasi cloud."""
        conn = self._get_connection()
        cursor = conn.cursor()
        cursor.execute('''
            INSERT OR REPLACE INTO log_laporan (tanggal, baris, status, keterangan, is_synced, waktu_eksekusi)
            VALUES (?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
        ''', (tanggal, baris, status, keterangan, is_synced))
        row_id = cursor.lastrowid
        conn.commit()
        conn.close()
        return row_id

    def ambil_log_pending_sync(self):
        """Mengambil daftar log yang belum berhasil sinkron ke cloud."""
        conn = self._get_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT id FROM log_laporan WHERE is_synced = 0")
        rows = cursor.fetchall()
        conn.close()
        return [r[0] for r in rows]

    def tandai_log_tersinkron(self, ids: list):
        """Menandai log tertentu sebagai sudah sukses sinkron."""
        if not ids:
            return
        conn = self._get_connection()
        cursor = conn.cursor()
        placeholders = ','.join(['?'] * len(ids))
        query = "UPDATE log_laporan SET is_synced = 1 WHERE id IN (" + placeholders + ")" # nosec B608
        cursor.execute(query, ids)
        conn.commit()
        conn.close()

    def ambil_laporan_tanggal(self, tanggal):
        conn = self._get_connection()
        query = "SELECT baris, status, keterangan, waktu_eksekusi FROM log_laporan WHERE tanggal = ? ORDER BY baris ASC"
        df = pd.read_sql_query(query, conn, params=(tanggal,))
        conn.close()
        return df

    def ambil_rekap_semua_tanggal(self, limit: Optional[int] = None):
        conn = self._get_connection()
        # Menggunakan MAX(waktu_eksekusi) agar baris paling baru diinput selalu di atas (Top 10)
        query = """
            SELECT 
                tanggal,
                COUNT(CASE WHEN status = 'SUKSES' THEN 1 END) as sukses,
                COUNT(CASE WHEN status = 'KUNING' THEN 1 END) as inkomplet,
                COUNT(CASE WHEN status = 'MERAH' THEN 1 END) as gagal_data,
                COUNT(CASE WHEN status = 'ERROR' THEN 1 END) as gagal_sistem,
                COUNT(*) as total,
                MAX(waktu_eksekusi) as last_exec
            FROM log_laporan
            GROUP BY tanggal
            ORDER BY last_exec DESC
        """
        
        # C-015: Parameterized LIMIT untuk mencegah SQL Injection
        params = []
        if limit and isinstance(limit, int):
            query += " LIMIT ?"
            params.append(limit)
            
        try:
            df = pd.read_sql_query(query, conn, params=params)
        except Exception as e:
            logger.error(f"Gagal ambil rekap tanggal: {e}")
            df = pd.DataFrame(columns=['tanggal', 'sukses', 'inkomplet', 'gagal_data', 'gagal_sistem', 'total', 'last_exec'])
        conn.close()
        return df

    def ambil_rekap_tanggal_spesifik(self, tanggal: str):
        """Mendukung fitur Fokus Pencarian (menampilkan 1 baris saja)."""
        conn = self._get_connection()
        query = """
            SELECT 
                tanggal,
                COUNT(CASE WHEN status = 'SUKSES' THEN 1 END) as sukses,
                COUNT(CASE WHEN status = 'KUNING' THEN 1 END) as inkomplet,
                COUNT(CASE WHEN status = 'MERAH' THEN 1 END) as gagal_data,
                COUNT(CASE WHEN status = 'ERROR' THEN 1 END) as gagal_sistem,
                COUNT(*) as total
            FROM log_laporan
            WHERE tanggal = ?
            GROUP BY tanggal
        """
        try:
            df = pd.read_sql_query(query, conn, params=(tanggal,))
        except Exception:
            df = pd.DataFrame(columns=['tanggal', 'sukses', 'inkomplet', 'gagal_data', 'gagal_sistem', 'total'])
        conn.close()
        return df

    def export_laporan_excel(self, tanggal, folder_tujuan):
        df = self.ambil_laporan_tanggal(tanggal)
        if df.empty:
            return False, "Tidak ada data laporan pada tanggal tersebut."

        nama_file = f"Laporan_Nextflow_{tanggal}.xlsx"
        path_output = os.path.join(folder_tujuan, nama_file)

        df = df.rename(columns={
            "baris": "Baris Excel Asli",
            "status": "Status Akhir",
            "keterangan": "Keterangan Detail",
            "waktu_eksekusi": "Waktu Eksekusi"
        })

        df.to_excel(path_output, index=False, engine='openpyxl')

        try:
            from core.config import ExcelColor
            wb = load_workbook(path_output)
            ws = wb.active
            warna_map = {
                "SUKSES": ExcelColor.HIJAU,
                "KUNING": ExcelColor.KUNING,
                "MERAH": ExcelColor.MERAH,
                "ERROR": ExcelColor.BIRU,
            }
            for baris_excel in range(2, len(df) + 2):
                status_cell = ws.cell(row=baris_excel, column=2).value
                if status_cell in warna_map:
                    hex_color = warna_map[status_cell]
                    fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                    for col in range(1, 5):
                        ws.cell(row=baris_excel, column=col).fill = fill
            wb.save(path_output)
            return True, path_output
        except Exception as e:
            return False, f"Gagal mewarnai laporan: {e}"
